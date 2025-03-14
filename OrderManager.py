from openpyxl import Workbook, load_workbook
from datetime import datetime
from PyQt6.QtWidgets import  QMessageBox, QTableWidgetItem
import os


class OrderManager:
    def __init__(self, table_widget, total_label, discount_input):
        self.table_widget = table_widget  # Bảng hóa đơn (QTableWidget)
        self.total_label = total_label  # Tổng tiền (QLineEdit hoặc QLabel)
        self.discount_input = discount_input  # Giảm giá (QLineEdit)
        self.order_items = {}  # Từ điển chứa hóa đơn: {name: {"price": float, "quantity": int}}


    def save_to_excel(self, table_name, current_user):
        """Lưu hóa đơn vào file hoadon.xlsx, chỉ thêm tiêu đề nếu file chưa có"""
        if not table_name or not current_user:
            QMessageBox.warning(None, "Lỗi", "Vui lòng chọn bàn và đảm bảo nhân viên đã đăng nhập!")
            return

        if not self.order_items:
            QMessageBox.warning(None, "Lỗi", "Không có món nào trong hóa đơn!")
            return

        try:
            file_name = "hoadon.xlsx"

            # Kiểm tra xem file đã tồn tại chưa
            if os.path.exists(file_name):
                wb = load_workbook(file_name)
                ws = wb.active
                # Nếu file không có dữ liệu, thêm tiêu đề
                if ws.max_row == 0:
                    ws.append(["Bàn", "Nhân viên", "Ngày giờ", "Tên món", "Số lượng", "Đơn giá", "Thành tiền"])
            else:
                # Nếu file chưa tồn tại, tạo mới và thêm tiêu đề
                wb = Workbook()
                ws = wb.active
                ws.title = "Danh sách hóa đơn"
                ws.append(["Bàn", "Nhân viên", "Ngày giờ", "Tên món", "Số lượng", "Đơn giá", "Thành tiền"])

            # Ghi dữ liệu hóa đơn vào file
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            total_amount = 0
            for name, details in self.order_items.items():
                quantity = details["quantity"]
                price = details["price"]
                total_price = quantity * price
                total_amount += total_price
                ws.append([table_name, current_user, current_time, name, quantity, f"{price:,.0f} VND",
                           f"{total_price:,.0f} VND"])

            # Thêm dòng tổng cộng
            ws.append(["", "", "", "", "Tổng cộng:", "", f"{total_amount:,.0f} VND"])

            # Lưu file Excel
            wb.save(file_name)
            QMessageBox.information(None, "Thành công", "Hóa đơn đã được lưu thành công!")
        except Exception as e:
            QMessageBox.critical(None, "Lỗi", f"Không thể lưu hóa đơn: {str(e)}")

    def add_to_order(self, name, price):
        """Thêm món vào hóa đơn hoặc cập nhật số lượng nếu món đã tồn tại"""
        if name in self.order_items:
            # Nếu món đã tồn tại, tăng số lượng
            self.order_items[name]["quantity"] += 1
        else:
            # Nếu món chưa tồn tại, thêm vào với số lượng ban đầu là 1
            self.order_items[name] = {"price": price, "quantity": 1}

        # Cập nhật lại bảng hóa đơn
        self.update_order_table()

    def update_order_table(self):
        """Cập nhật bảng hóa đơn và tổng tiền"""
        self.table_widget.setRowCount(0)  # Xóa dữ liệu cũ trong bảng

        for name, details in self.order_items.items():
            row_position = self.table_widget.rowCount()
            self.table_widget.insertRow(row_position)
            self.table_widget.setItem(row_position, 0, QTableWidgetItem(name))  # Tên món
            self.table_widget.setItem(row_position, 1, QTableWidgetItem(str(details["quantity"])))  # Số lượng
            self.table_widget.setItem(row_position, 2, QTableWidgetItem(f"{details['price']:,.0f} VND"))  # Giá mỗi món
            total_price = details["price"] * details["quantity"]
            self.table_widget.setItem(row_position, 3,
                                      QTableWidgetItem(f"{total_price:,.0f} VND"))  # Tổng tiền từng món

        # Cập nhật tổng tiền
        self.update_total()

    def update_total(self):
        """Tính tổng tiền và hiển thị"""
        total = sum(details["price"] * details["quantity"] for details in self.order_items.values())
        discount = self.get_discount()  # Lấy giá trị giảm giá
        total_after_discount = max(0, total - discount)  # Đảm bảo tổng tiền không âm

        # Hiển thị tổng tiền
        self.total_label.setText(f"{total_after_discount:,.0f} VND")

    def get_discount(self):
        """Lấy giá trị giảm giá từ input"""
        try:
            discount_text = self.discount_input.text()
            if not discount_text:
                return 0  # Không có giảm giá
            return float(discount_text)  # Trả về giá trị giảm giá
        except ValueError:
            QMessageBox.warning(None, "Lỗi", "Giảm giá không hợp lệ!")
            return 0

    def apply_discount(self):
        """Áp dụng giảm giá và cập nhật lại tổng tiền"""
        self.update_total()  # Cập nhật tổng tiền với giảm giá

    def remove_item(self, row):
        """Xóa món khỏi hóa đơn"""
        if row < 0 or row >= self.table_widget.rowCount():
            QMessageBox.warning(None, "Lỗi", "Vui lòng chọn một món để xóa!")
            return

        # Lấy tên món từ dòng được chọn
        item_name = self.table_widget.item(row, 0).text()
        if item_name in self.order_items:
            del self.order_items[item_name]  # Xóa món khỏi từ điển

        # Cập nhật lại bảng hóa đơn
        self.update_order_table()

    def clear_order(self):
        """Xóa toàn bộ hóa đơn"""
        self.order_items = {}
        self.table_widget.setRowCount(0)
        self.update_total()

    def get_order(self):
        """Trả về danh sách hóa đơn hiện tại"""
        return [
            {
                "name": name,
                "quantity": details["quantity"],
                "price": details["price"],
                "total": details["price"] * details["quantity"]
            }
            for name, details in self.order_items.items()]
